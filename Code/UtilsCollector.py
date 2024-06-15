import pandas as pd
import numpy as np
from termcolor import colored
import re
from sklearn.preprocessing import StandardScaler
import random
import openpyxl
import scipy.optimize as opt

class UtilsCollector:
    def __init__(self, exp_params):
        self.exp_params = exp_params
        self.app_indices = []

    def no_nan_checks(self, x):
        if type(x) is np.ndarray:
            if np.isnan(x).any():
                return False
        return True

    def load_observations_GP(self, filename, sheetname):
        df_all_cols = pd.read_excel(filename, sheet_name=sheetname)
        df_y_pref_cols = df_all_cols.iloc[:, -1]
        x_scaled = None

        if sheetname == self.exp_params.obj_sheet or sheetname == self.exp_params.obj_sheet\
                or sheetname == self.exp_params.baseline_sheetname:

            df_x_cols = df_all_cols[self.exp_params.input_variables]
            x_np = df_x_cols.to_numpy()
            x_min = self.exp_params.bounds[:, 0]
            x_max = self.exp_params.bounds[:, 1]
            x_scaled = (x_np - x_min) / (x_max - x_min)

            y_pref_np_unscaled = df_y_pref_cols.to_numpy().reshape(-1, 1)

            # scaler = StandardScaler().fit(y_pref_np_unscaled)
            # y_pref_np = scaler.transform(y_pref_np_unscaled)

            y_pref_np = y_pref_np_unscaled.copy()

            if not self.no_nan_checks(df_x_cols):
                print(colored("Nan values detected in the observation file. Please fix them and restart", "red"))
                exit(0)
        else:
            y_pref_np = self.obtain_preferences_array(df_y_pref_cols.to_numpy())

        return x_scaled, y_pref_np

    def obtain_preferences_array(self, pref):

        pref_np = np.empty(shape=(len(pref), 2), dtype=np.int)
        for i in range(len(pref)):
            match = re.search("^([0-9]+),([0-9]+)$", pref[i])
            if not match:
                print(colored("Invalid format encountered for preferences", "red"))
                exit(0)
            pref_np[i][0] = int(match.group(1))
            pref_np[i][1] = int(match.group(2))

        return pref_np

    def conduct_init_property_evaluation(self, x):

        pref_prop1_all = []
        pref_prop2_all = []

        if self.exp_params.num_features == 1:
            for i in range(len(x)):
                for j in range(i, len(x)):
                    if i != j:
                        if self.exp_params.prop1(x[i, :]) > self.exp_params.prop1(x[j, :]):
                            pref_prop1_all.append(str(i) + "," + str(j))
                        else:
                            pref_prop1_all.append(str(j) + "," + str(i))

                        if self.exp_params.prop2(x[i, :]) > self.exp_params.prop2(x[j, :]):
                            pref_prop2_all.append(str(i) + "," + str(j))
                        else:
                            pref_prop2_all.append(str(j) + "," + str(i))

            rand_indices_prop1 = np.random.randint(low=0, high=len(pref_prop1_all), size=(len(pref_prop1_all) - 3))
            rand_indices_prop1.sort()
            pref_prop1 = [pref_prop1_all[i] for i in rand_indices_prop1]

            rand_indices_prop2 = np.random.randint(low=0, high=len(pref_prop2_all), size=(len(pref_prop2_all) - 3))
            rand_indices_prop2.sort()
            pref_prop2 = [pref_prop2_all[j] for j in rand_indices_prop2]

            rand_indices = [rand_indices_prop1, rand_indices_prop2]
            return pref_prop1, pref_prop2

        elif self.exp_params.num_features > 1:
            for i in range(len(x)):
                for j in range(i, len(x)):
                    if i != j:
                        if self.exp_params.prop1(np.array([x[i, :]])) > self.exp_params.prop1(np.array([x[j, :]])):
                            pref_prop1_all.append(str(i) + "," + str(j))
                        else:
                            pref_prop1_all.append(str(j) + "," + str(i))

                        if self.exp_params.prop2(np.array([x[i, :]])) > self.exp_params.prop2(np.array([x[j, :]])):
                            pref_prop2_all.append(str(i) + "," + str(j))
                        else:
                            pref_prop2_all.append(str(j) + "," + str(i))

            rand_indices_prop1 = np.random.randint(low=0, high=len(pref_prop1_all), size=(len(pref_prop1_all) - 3))
            rand_indices_prop1.sort()
            pref_prop1 = [pref_prop1_all[i] for i in rand_indices_prop1]

            rand_indices_prop2 = np.random.randint(low=0, high=len(pref_prop2_all), size=(len(pref_prop2_all) - 3))
            rand_indices_prop2.sort()
            pref_prop2 = [pref_prop2_all[j] for j in rand_indices_prop2]

            rand_indices = [rand_indices_prop1, rand_indices_prop2]
            return pref_prop1, pref_prop2

    def conduct_experiment(self, x):

        if self.exp_params.true_func == "Benchmark":

            # # exp{-(x-2)^2} + exp{-((x-6)^2)/10} + (1/(X^2 +1))
            self.exp_params.prop1 = lambda x: np.exp(-(x - 2) ** 2)
            self.exp_params.prop2 = lambda x: (1 / (x ** 2 + 1))
            return np.exp(-(x - 2) ** 2) + np.exp(-(x - 6) ** 2 / 10) + (1 / (x ** 2 + 1))

            # oscillator
            # exp(-x)sin(3.pi.x) + 01
            # return (np.exp(-x) * np.sin(3 * np.pi * x)) + 1

            # Complicated Oscillator circuit
            # val = 0
            # if x < 10:
            #     val = (np.exp(-x) * np.sin(1.5 * np.pi * x)) + 1
            # elif x > 10 and x <= 20:
            #     x = x - 10
            #     val = (np.exp(-x) * np.sin(1.5 * np.pi * x)) + 1
            # elif x > 20 and x <= 30:
            #     x = x - 20
            #     val = (np.exp(-x) * np.sin(1.5 * np.pi * x)) + 1
            # return val

            # exp(-x)sin(8.pi.x) + 1
            # return (np.exp(-x) * np.sin(8 * np.pi * x)) + 1

            # Benchmark Function exp(-x)*sin(2.pi.x)(maxima), -exp(-x)*sin(2.pi.x) (minima)
            # return (np.exp(-x) * np.sin(2 * np.pi * x))

            # Gramacy and Lee function sin(10.pi.x/2x)+(x-1)^4; minima = -2.874 @ x=0.144; -sin(10.pi.x/2x)-x-1)^4; maxima = 2.874 @x=0.144
            # return -1 * (((np.sin(10 * np.pi * x))/(2*x)) + (x-1) ** 4)

            # Levy function w = 1+(x-1)/4  y = (sin(w*pi))^2 + (w-1)^2(1+(sin(2w*pi))^2) max =0
            # w = -0.5+((x-1)/4)
            # w = 1+((x-1)/4)
            # value = ((np.sin(w * np.pi))**2 + ((w-1)**2)*(1+((np.sin(2*w*np.pi)) ** 2 )))
            # return -1 * value

            # square wave function
            # y = np.array([])
            #
            # for each_x in x:
            #     each_y = np.sin(2 * np.pi * each_x)
            #     if each_y < 0:
            #         each_y = -1
            #     elif each_y > 0:
            #         each_y = 1
            #     else:
            #         each_y = 0
            #     y = np.append(y, each_y)
            # return y.reshape(-1, 1)

            # Triangular wave function
            # return (2 * np.arcsin(np.sin(np.pi * x))) / (np.pi)

            # Chirpwave function
            # y = np.array([])
            # f = 1
            # for each_x in x:
            #     if each_x < 8:
            #         f = 0.35
            #     elif each_x > 8 and each_x <= 15:
            #         f = 1.25
            #     elif each_x > 15 and each_x <= 20:
            #         f = 0.35
            #     val = np.sin(2 * np.pi * f * each_x)
            #     y = np.append(y, val)
            # return y.reshape(-1, 1)

            # Sinc Function
            # return np.sinc(x - 10) + np.sinc(x) + np.sinc(x + 10)

            # Gaussian Mixtures
            # y = np.array([])
            # for each_x in x:
            #     if each_x <= 5:
            #         sig = 0.4
            #         mean = 2.5
            #     elif each_x > 5 and each_x <= 10:
            #         sig = 0.7
            #         mean = 7.5
            #     elif each_x > 10 and each_x <= 15:
            #         sig = 0.6
            #         mean= 12.5
            #     val = 1 / (sig * np.sqrt(2 * np.pi)) * np.exp(-0.5 * ((each_x - mean) / sig) * (each_x - mean) / sig)
            #     y = np.append(y, val)
            # return y.reshape(-1, 1)

            # Linear function
            # return 0.1 * x + 0.2

            # Linear Sin Function
            # return 0.7*x + 1 + np.sin(2*np.pi*x)

        elif self.exp_params.true_func == 'branin2d':
            # branin 2d fucntion
            # a = 1, b = 5.1 ⁄ (4π2), c = 5 ⁄ π, r = 6, s = 10 and t = 1 ⁄ (8π)
            # y = a * (x2 - b * x1 **2 + c * x1 - r) ** 2 + s * (1 - t) * cos(x1) + s
            x1 = x[:, 0]
            x2 = x[:, 1]
            a = 1;
            b = 5.1 / (4 * (np.pi ** 2));
            c = 5 / np.pi;
            r = 6;
            s = 10;
            t = 1 / (8 * np.pi)
            value = a * (x2 - b * x1 ** 2 + c * x1 - r) ** 2 + s * (1 - t) * np.cos(x1) + s
            value = -1 * value.reshape((-1, 1))
            return value

        elif self.exp_params.true_func == 'sphere':
            # simple sphere equation
            # Z = X**2 + Y**2
            x1 = x[:, 0]
            x2 = x[:, 1]
            value = (x1 ** 2 + x2 ** 2)
            # Regression setting
            # value = -1 * value.reshape(-1, 1)
            value = 1 * value.reshape(-1, 1)
            return value

        elif self.exp_params.true_func == 'hartmann3d':

            alpha = np.array([1.0, 1.2, 3.0, 3.2])
            A_array = [[3, 10, 30],
                       [0.1, 10, 35],
                       [3, 10, 30],
                       [0.1, 10, 35]
                       ]
            A = np.matrix(A_array)

            P_array = [[3689, 1170, 2673],
                       [4699, 4387, 7470],
                       [1091, 8732, 5547],
                       [381, 5743, 8828]
                       ]

            P = np.matrix(P_array)
            P = P * 1e-4

            sum = 0
            for i in np.arange(0, 4):
                alpha_value = alpha[i]
                inner_sum = 0
                for j in np.arange(0, 3):
                    inner_sum += A.item(i, j) * ((x[:, j] - P.item(i, j)) ** 2)
                sum += alpha_value * np.exp(-1 * inner_sum)
            # extra -1 is because we are finding maxima instead of the minima f(-x)
            value = (-1 * -1 * sum).reshape(-1, 1)
            return value

        elif self.exp_params.true_func == 'hartmann6d':

            alpha = np.array([1.0, 1.2, 3.0, 3.2])
            A_array = [[10, 3, 17, 3.50, 1.7, 8],
                       [0.05, 10, 17, 0.10, 8, 14],
                       [3, 3.5, 1.7, 10, 17, 8],
                       [17, 8, 0.05, 10, 0.1, 14]
                       ]
            A = np.matrix(A_array)

            P_array = [[1312, 1696, 5569, 124, 8283, 5886],
                       [2329, 4135, 8307, 3736, 1004, 9991],
                       [2348, 1451, 3522, 2883, 3047, 6650],
                       [4047, 8828, 8732, 5743, 1091, 381]
                       ]
            P = np.matrix(P_array)
            P = P * 1e-4

            sum = 0
            for i in np.arange(0, 4):
                alpha_value = alpha[i]
                inner_sum = 0
                for j in np.arange(0, 6):
                    inner_sum += A.item(i, j) * ((x[:, j] - P.item(i, j)) ** 2)
                sum += alpha_value * np.exp(-1 * inner_sum)
            # extra -1 is because we are finding maxima instead of the minima f(-x)
            value = (-1 * -1 * sum).reshape(-1, 1)
            return value

        elif self.exp_params.true_func == "syn2d":
            # # Custom synthetic function exp(-x)*sin(2*pi*x)converted from 1d to 2d max =0.62198323 @x=0.224,0.223
            x1 = x[:, 0]
            x2 = x[:, 1]
            value = (np.exp(-x1) * np.sin(2 * np.pi * x1)) * (np.exp(-x2) * np.sin(2 * np.pi * x2))
            value = value.reshape((-1, 1))
            return value

        elif self.exp_params.true_func == "levy2d":
            # # Levy 2d function minima is 0 at X = (1,1)

            X1 = x[:, 0]
            X2 = x[:, 1]

            w1 = 1 + ((X1 - 1) / 4)
            w2 = 1 + ((X2 - 1) / 4)

            value = ((np.sin(np.pi * w1)) ** 2) + ((w1 - 1) ** 2) * (1 + 10 * ((np.sin((np.pi * w1) + 1)) ** 2)) + (
                    (w2 - 1) ** 2) * (
                            1 + ((np.sin(2 * np.pi * w2)) ** 2))

            value = (-1 * value).reshape((-1, 1))
            return value

        elif self.exp_params.true_func == "ackley2d":
            # # Ackley 2d function minima is 0 at X = (0,0)

            X1 = x[:, 0]
            X2 = x[:, 1]

            a = 20
            b = 0.2
            c = 2 * np.pi
            value = (-20 * np.exp(-0.2 * np.sqrt(0.5 * (X1 ** 2 + X2 ** 2))) - np.exp(
                0.5 * (np.cos(2 * np.pi * X1) + np.cos(2 * np.pi * X2))) + 20 + np.exp(1))

            value = (-1 * value).reshape((-1, 1))
            return value

        elif self.exp_params.true_func == "egg2d":
            # # egg holder 2d function  Maxima = 959.64008971 @ x =[512,404.25425425]
            X1 = x[:, 0]
            X2 = x[:, 1]

            value = (-(X2 + 47) * np.sin(np.sqrt(np.abs(X2 + (X1 / 2) + 47))) - X1 * np.sin(
                np.sqrt(np.abs(X1 - (X2 + 47)))))

            value = (-1 * value).reshape((-1, 1))
            return value

        elif self.exp_params.true_func == "michalewicz2d":
            # Michalewicz maxima = 1.80116404 @ x = [2.20446091 1.56922396]
            X1 = x[:, 0]
            X2 = x[:, 1]

            value = np.sin(X1) * ((np.sin(((X1 ** 2) / np.pi))) ** 20) + np.sin(X2) * (
                    (np.sin((2 * (X2 ** 2) / np.pi))) ** 20)

            value = (1 * value).reshape((-1, 1))
            return value

        # Ackley 1d
        elif self.exp_params.true_func == "Ackley1D":
            self.exp_params.prop1 = lambda x: x ** 2
            self.exp_params.prop2 = lambda x: np.cos(x)
            return 20 * np.exp(-0.2 * (x ** 2)) + np.exp(np.cos(2 * np.pi * x)) - 20 - np.exp(1)

        # Matyas 2D
        elif self.exp_params.true_func == "Matyas2D":
            self.exp_params.prop1 = lambda x: x[:, 0] ** 2 + x[:, 1] ** 2
            self.exp_params.prop2 = lambda x: x[:, 0] * x[:, 1]
            return 0.26 * (x[:, 0] ** 2 + x[:, 1] ** 2) - 0.48 * (x[:, 0] * x[:, 1])

            # Griewank 3D
        elif self.exp_params.true_func == "Griewank3D":
            self.exp_params.prop1 = lambda x: x[:, 0] ** 2 + x[:, 1] ** 2 + x[:, 2] ** 2
            self.exp_params.prop2 = lambda x: (np.cos(x[:, 0])) * (np.cos(x[:, 1]/np.sqrt(2))) * (np.cos(x[:, 2]/np.sqrt(3)))
            return (-1 * (x[:, 0] ** 2 + x[:, 1] ** 2 + x[:, 2] ** 2)/4000) + ((np.cos(x[:, 0])) *
                                                                                (np.cos(x[:, 1]/np.sqrt(2))) *
                                                                                (np.cos(x[:, 2]/np.sqrt(3)))) -1

        elif self.exp_params.true_func == "Griewank5D":
            self.exp_params.prop1 = lambda x: x[:, 0] ** 2 + x[:, 1] ** 2 + x[:, 2] ** 2 + x[:, 3] ** 2 + x[:, 4] ** 2
            self.exp_params.prop2 = lambda x: (np.cos(x[:, 0])) * (np.cos(x[:, 1]/np.sqrt(2))) * \
                                              (np.cos(x[:, 2]/np.sqrt(3))) * (np.cos(x[:, 3]/np.sqrt(4))) * \
                                              (np.cos(x[:, 4] / np.sqrt(5)))
            return (-1 * (x[:, 0] ** 2 + x[:, 1] ** 2 + x[:, 2] ** 2 + x[:, 3] ** 2 + x[:, 4] ** 2)/4000) + \
                   ((np.cos(x[:, 0])) * (np.cos(x[:, 1]/np.sqrt(2))) * (np.cos(x[:, 2]/np.sqrt(3)))
                    * (np.cos(x[:, 3]/np.sqrt(4))) * (np.cos(x[:, 4]/np.sqrt(5)))) - 1

        elif self.exp_params.true_func == "Rosenbrock3D":
            self.exp_params.prop1 = lambda x: -100 * (x[:, 1] - (x[:, 0])**2)**2 - 100 * (x[:, 2] - (x[:, 1])**2)**2
            self.exp_params.prop2 = lambda x: -(x[:, 0] - 1)**2 - (x[:, 1] - 1)**2
            return -100 * (x[:, 1] - (x[:, 0])**2)**2 - (x[:, 0] - 1)**2 - 100 * (x[:, 2] - (x[:, 1])**2)**2 - (x[:, 1]
                                                                                                                - 1)**2

    def generate_initial_pref(self, x, y):

        pref_h_all = []
        # generate preference for human objective
        for i in range(len(x)):
            for j in range(i + 1, len(x)):
                if y[i] > y[j]:
                    pref_h_all.append(str(i) + "," + str(j))
                else:
                    pref_h_all.append(str(j) + "," + str(i))

        rand_indices_hum = np.random.randint(low=0, high=len(pref_h_all), size=(len(pref_h_all)-3))
        rand_indices_hum.sort()
        pref_h = [pref_h_all[i] for i in rand_indices_hum]

        # generate preference for property 1 and property 2
        pref_prop1, pref_prop2 = self.conduct_init_property_evaluation(x)

        return pref_h, pref_prop1, pref_prop2

    def generate_initial_observations(self):

        # Generate random points
        random_points = []
        X_init = []

        # Generate specified (number of unseen data points) random numbers for each dimension
        for dim in np.arange(self.exp_params.num_features):
            random_data_point_each_dim = np.random.uniform(self.exp_params.bounds[dim][0],
                                                           self.exp_params.bounds[dim][1],
                                                           self.exp_params.num_init_obs). \
                reshape(1, self.exp_params.num_init_obs)
            random_points.append(random_data_point_each_dim)
        random_points = np.vstack(random_points)

        # Generated values are to be reshaped into input points in the form of x1=[x11, x12, x13].T
        for sample_num in np.arange(self.exp_params.num_init_obs):
            array = []
            for dim_count in np.arange(self.exp_params.num_features):
                array.append(random_points[dim_count, sample_num])
            X_init.append(array)
        X_init = np.vstack(X_init)

        Y_init = self.conduct_experiment(X_init)
        pref_prop1, pref_prop2 = self.conduct_init_property_evaluation(X_init)

        X = np.around(X_init, decimals=3)
        Y = np.around(Y_init, decimals=3).reshape(-1, 1)

        writer = pd.ExcelWriter(self.exp_params.obs_file)

        source_hf = ['Initial'] * self.exp_params.num_init_obs
        source_hf = np.array([source_hf]).reshape(-1, 1)

        hf_init_observations = np.hstack((X, Y))
        hf_init_file_contents = np.hstack((source_hf, hf_init_observations))
        df_hf_exp = pd.DataFrame(hf_init_file_contents,
                              columns=["Source"] + self.exp_params.input_variables + self.exp_params.output_variable)
        df_hf_exp.to_excel(writer, sheet_name=self.exp_params.obj_sheet)

        baseline_writer = pd.ExcelWriter(self.exp_params.baseline_obs_file)
        df_hf_exp.to_excel(baseline_writer, sheet_name=self.exp_params.baseline_sheetname)
        baseline_writer.save()

        source_prop1 = ['Initial'] * len(pref_prop1)
        source_prop1 = np.array([source_prop1]).reshape(-1, 1)
        prop1_file_contents = np.hstack((source_prop1, np.array(pref_prop1).reshape(-1, 1)))
        df_prop1 = pd.DataFrame(prop1_file_contents, columns=["Source"] + self.exp_params.pref_variable)
        df_prop1.to_excel(writer, sheet_name=self.exp_params.props_sheetname[0])

        source_prop2 = ['Initial'] * len(pref_prop2)
        source_prop2 = np.array([source_prop2]).reshape(-1, 1)
        prop2_file_contents = np.hstack((source_prop2, np.array(pref_prop2).reshape(-1, 1)))
        df_prop2 = pd.DataFrame(prop2_file_contents, columns=["Source"] + self.exp_params.pref_variable)
        df_prop2.to_excel(writer, sheet_name=self.exp_params.props_sheetname[1])

        df_empty = pd.DataFrame()
        df_empty.to_excel(writer, sheet_name=self.exp_params.arms_sheetname)

        writer.save()

    def calculate_arm_reward_GO(self, arm_type, arm_gp_obj):

        print(colored("Calculate reward for arm:{} using Global Optimiser".format(arm_type), "green"))

        print("Computing maximum likelihood for arm type:", arm_type)
        max_likelihood = arm_gp_obj.compute_max_loglikelihood()

        # Finding the maxima of the posterior
        post_max_value = None
        post_max = - 1 * float("inf")

        # Data structure to create the starting points for the scipy.minimize method
        random_points = []
        starting_points = []

        # Depending on the number of dimensions and bounds, generate random multi-starting points to find maxima
        for dim in np.arange(arm_gp_obj.num_features):
            random_data_point_each_dim = np.random.uniform(0, 1, arm_gp_obj.number_of_restarts_likelihood). \
                reshape(1, arm_gp_obj.number_of_restarts_likelihood)
            random_points.append(random_data_point_each_dim)

        # Vertically stack the arrays of randomly generated starting points as a matrix
        random_points = np.vstack(random_points)

        # Reformat the generated random starting points in the form [x1 x2].T for the specified number of restarts
        for sample_num in np.arange(arm_gp_obj.number_of_restarts_likelihood):
            array = []
            for dim_count in np.arange(arm_gp_obj.num_features):
                array.append(random_points[dim_count, sample_num])
            starting_points.append(array)
        starting_points = np.vstack(starting_points)

        total_bounds = arm_gp_obj.len_scale_bounds.copy()

        for ind in np.arange(arm_gp_obj.number_of_restarts_likelihood):

            init_point = starting_points[ind]

            # print("Initial length scale: ", init_len_scale)
            maxima = opt.minimize(lambda x: -self.compute_posterior_for_x(x, arm_gp_obj),
                                  init_point,
                                  method='L-BFGS-B',
                                  tol=0.01,
                                  options={'maxfun': 20, 'maxiter': 20},
                                  bounds=total_bounds)

            post_temp = maxima['x'][:arm_gp_obj.num_features]
            post_max_new = self.compute_posterior_for_x(post_temp, arm_gp_obj)

            if post_max_new > post_max:
                print("New maximum for posterior: ", post_max_new, " found for X= ",
                      maxima['x'][: arm_gp_obj.num_features])
                post_max_value = maxima
                post_max = post_max_new

        print(colored("Global optima of the posterior sample:{} is found at: {}".format(post_max,
                                                                    post_max_value['x'][: arm_gp_obj.num_features])))
        return max_likelihood, post_max_value['x'][: arm_gp_obj.num_features-self.exp_params.num_properties]


    def compute_posterior_for_x(self, Xs, arm_gp_obj):

        Xs = np.array([Xs])
        K_xs_xs = arm_gp_obj.compute_kernel(Xs, Xs, arm_gp_obj.char_len_scale)

        # Sample 3 standard normals for each of the unseen data points
        standard_normals = np.random.normal(size=(1, 1))

        # Compute mean and variance
        mean, variance, factor1 = arm_gp_obj.compute_mean_var(Xs, arm_gp_obj.X, arm_gp_obj.y_fmap)

        # compute posteriors for the data points
        newL = np.linalg.cholesky(K_xs_xs + 1e-6 * np.eye(1) - np.dot(factor1.T, factor1))
        f_post = mean.reshape(-1, 1) + np.dot(newL, standard_normals)
        return f_post


    def calculate_arm_reward(self, type, arm_gp_obj):

        grid = np.linspace(0, 1, self.exp_params.num_testpoints)
        if type == "exp":
            Xs = grid.reshape(-1, 1)
        elif type == "human":
            Xs = []
            Xs.append(grid)
            Xs.append(grid)
            Xs.append(grid)
            Xs = np.array(Xs).T

        # compute the covariances between the unseen data points i.e K**
        K_xs_xs = arm_gp_obj.compute_kernel(Xs, Xs,  arm_gp_obj.char_len_scale)

        # Sample 3 standard normals for each of the unseen data points
        standard_normals = np.random.normal(size=(self.exp_params.num_testpoints, 1))

        # Compute mean and variance
        mean, variance, factor1 = arm_gp_obj.compute_mean_var(Xs, arm_gp_obj.X, arm_gp_obj.y_fmap)

        # compute posteriors for the data points
        newL = np.linalg.cholesky(
            K_xs_xs + 1e-6 * np.eye(self.exp_params.num_testpoints) - np.dot(factor1.T, factor1))
        f_post = mean.reshape(-1, 1) + np.dot(newL, standard_normals)

        max_index = np.argmax(f_post)
        max_reward = np.max(f_post)
        x_maxima = Xs[max_index, 0]

        return max_reward, x_maxima

    def generate_auxillary_inputs(self, X_exp, prop1_gp_obj, prop2_gp_obj):

        X_exp_aux = []
        for each_x in X_exp:
            x_prop1, _ = prop1_gp_obj.gp_predict(np.array([each_x]))
            x_prop2, _ = prop2_gp_obj.gp_predict(np.array([each_x]))
            X_exp_aux.append(np.append(each_x, [x_prop1[0][0], x_prop2[0][0]]))
        X_exp_aux = np.vstack(X_exp_aux)
        return X_exp_aux

    def update_preferences(self, x_new, y_new):

        # Load previous observations from file
        df_all_cols = pd.read_excel(self.exp_params.obs_file, sheet_name=self.exp_params.obj_sheet)
        df_x_cols = df_all_cols[self.exp_params.input_variables]
        x_np = df_x_cols.to_numpy()
        df_y = df_all_cols[self.exp_params.output_variable]
        y_np = df_y.to_numpy()

        # # Human objective preferences
        # pref_h = []
        # rand_indices_hum = np.random.randint(low=0, high=len(y_np), size=(len(y_np)-3))
        # rand_indices_hum.sort()
        #
        # for each_idx in rand_indices_hum:
        #     if y_np[each_idx] > y_new:
        #         pref_h.append(str(each_idx)+","+str(len(y_np)))
        #     else:
        #         pref_h.append(str(len(y_np)) + "," + str(each_idx))
        #
        # source_hobj = ['BO'] * len(pref_h)
        # source_hobj = np.array([source_hobj]).reshape(-1, 1)
        #
        # workbook = openpyxl.load_workbook(filename=self.exp_params.obs_file)
        # hobj_sheet = workbook[self.exp_params.hum_obj_sheet]
        # xl_indices = np.arange(hobj_sheet.max_row-1, hobj_sheet.max_row-1+len(pref_h)).reshape(-1, 1)
        # hobj_contents = np.hstack((xl_indices, source_hobj, np.array(pref_h).reshape(-1, 1)))
        #
        # for each_row in hobj_contents:
        #     hobj_sheet.append(each_row.tolist())
        # workbook.save(self.exp_params.obs_file)

        # Property 1 Preferences
        pref_prop1 = []
        rand_indices_prop1 = np.random.randint(low=0, high=len(y_np), size=(len(y_np)-2))
        rand_indices_prop1.sort()

        for each_idx in rand_indices_prop1:
            if self.exp_params.prop1(np.array([x_np[each_idx]])) > self.exp_params.prop1(np.array([x_new])):
                pref_prop1.append(str(each_idx)+","+str(len(y_np)))
            else:
                pref_prop1.append(str(len(y_np)) + "," + str(each_idx))

        source_prop1 = ['BO'] * len(pref_prop1)
        source_prop1 = np.array([source_prop1]).reshape(-1, 1)

        workbook = openpyxl.load_workbook(filename=self.exp_params.obs_file)
        prop1_sheet = workbook[self.exp_params.props_sheetname[0]]
        xl_indices = np.arange(prop1_sheet.max_row-1, prop1_sheet.max_row-1 + len(pref_prop1)).reshape(-1, 1)
        prop1_contents = np.hstack((xl_indices, source_prop1, np.array(pref_prop1).reshape(-1, 1)))

        for each_row in prop1_contents:
            prop1_sheet.append(each_row.tolist())
        workbook.save(self.exp_params.obs_file)

        # Property 2 Preference
        pref_prop2 = []
        rand_indices_prop2 = np.random.randint(low=0, high=len(y_np), size=(len(y_np)-2))
        rand_indices_prop2.sort()

        for each_idx in rand_indices_prop2:
            if self.exp_params.prop2(np.array([x_np[each_idx]])) > self.exp_params.prop2(np.array([x_new])):
                pref_prop2.append(str(each_idx) + "," + str(len(y_np)))
            else:
                pref_prop2.append(str(len(y_np)) + "," + str(each_idx))

        source_prop2 = ['BO'] * len(pref_prop2)
        source_prop2 = np.array([source_prop2]).reshape(-1, 1)

        workbook = openpyxl.load_workbook(filename=self.exp_params.obs_file)
        prop2_sheet = workbook[self.exp_params.props_sheetname[1]]
        xl_indices = np.arange(prop2_sheet.max_row-1, prop2_sheet.max_row -1 + len(pref_prop2)).reshape(-1, 1)
        prop2_contents = np.hstack((xl_indices, source_prop2, np.array(pref_prop2).reshape(-1, 1)))

        for each_row in prop2_contents:
            prop2_sheet.append(each_row.tolist())
        workbook.save(self.exp_params.obs_file)

    # def load_observations_file(self, obs_filename, sheetname):
    #
    #     df_x_cols = pd.read_excel(obs_filename, sheet_name=sheetname, usecols=self.exp_params.x_colnames)
    #     df_y_cols = -1 * pd.read_excel(obs_filename, sheet_name=sheetname, usecols=self.exp_params.output_colnames)
    #     df_source = pd.read_excel(obs_filename, sheet_name=sheetname, usecols=self.exp_params.col_source)
    #     bo_source_indices = df_source.index[df_source["source"] == "BO"]
    #
    #     x = df_x_cols.copy()
    #     filt_num_x = x.filter(self.exp_params.num_colnames)
    #
    #     filtered_scaled_x, x_scaler = self.x_minmax_scaler_grid(filt_num_x.to_numpy())
    #     df_scaled_num_x = pd.DataFrame(data=filtered_scaled_x, columns=self.exp_params.num_colnames)
    #     y_scaler = StandardScaler().fit(df_y_cols.to_numpy())
    #     y_scaled = y_scaler.transform(df_y_cols.to_numpy())
    #
    #     for num_col in self.exp_params.num_colnames:
    #         df_x_cols[num_col] = df_scaled_num_x[num_col]
    #
    #     df_x_cols[self.exp_params.cat_colnames] = \
    #         df_x_cols[self.exp_params.cat_colnames].replace(self.exp_params.cat_lookup)
    #
    #     if not self.no_nan_checks(df_x_cols[self.exp_params.num_colnames]) or not self.no_nan_checks(y_scaled):
    #         print(colored("Nan values detected in the observation file. Please fix them and restart", "red"))
    #         exit(0)
    #
    #     return df_x_cols.to_numpy(), y_scaled, bo_source_indices, x_scaler
    #
    # def x_minmax_scaler_grid(self, x):
    #     search_grid = pd.read_pickle(self.exp_params.full_grid_file)
    #     x_scaler = MinMaxScaler()
    #     num_search_grid = search_grid.filter(self.exp_params.num_colnames)
    #     num_search_grid_np = num_search_grid.to_numpy()
    #     x_scaler.fit(num_search_grid_np)
    #     try:
    #         x_scaled = x_scaler.transform(x)
    #     except:
    #         print(colored("String values encountered while normalising X's in observations file", "red"))
    #         exit(0)
    #     return x_scaled, x_scaler
    #
    # def load_grid(self, grid_file, scale_type, x_scaler):
    #     if scale_type == "scale":
    #         full_grid = pd.read_pickle(grid_file)
    #         full_grid = full_grid.filter(self.exp_params.x_colnames)
    #         num_full_grid = full_grid.filter(self.exp_params.num_colnames)
    #         num_full_grid_np = num_full_grid.to_numpy()
    #         x_scaler.fit(num_full_grid_np)
    #         scaled_num_columns = x_scaler.transform(num_full_grid_np)
    #         df_scaled_num_columns = pd.DataFrame(data=scaled_num_columns, columns=self.exp_params.num_colnames)
    #
    #         for col in self.exp_params.num_colnames:
    #             full_grid[col] = df_scaled_num_columns[col]
    #
    #         full_grid[self.exp_params.cat_colnames] = \
    #             full_grid[self.exp_params.cat_colnames].replace(self.exp_params.cat_lookup)
    #
    #         return torch.tensor(full_grid.to_numpy(), device=device)
    #
    #     elif scale_type == "unscale":
    #         unsc_grid = pd.read_pickle(grid_file)
    #         return unsc_grid
    #
    # def check_if_duplicates(self, x_opt):
    #     df_x_cols = pd.read_excel(self.exp_params.obs_file, sheet_name=self.exp_params.obs_sheetname,
    #                               usecols=self.exp_params.input_colnames)
    #
    #     obs_row_index = df_x_cols[(df_x_cols["p1_mat"].values == x_opt["p1_mat"].values) &
    #                               (df_x_cols["p1_thick"].values == x_opt["p1_thick"].values) &
    #                               (df_x_cols["gap1"].values == x_opt["gap1"].values) &
    #                               (df_x_cols["p2_mat"].values == x_opt["p2_mat"].values) &
    #                               (df_x_cols["p2_thick"].values == x_opt["p2_thick"].values) &
    #                               (df_x_cols["gap2"].values == x_opt["gap2"].values) &
    #                               (df_x_cols["p3_mat"].values == x_opt["p3_mat"].values) &
    #                               (df_x_cols["p3_thick"].values == x_opt["p3_thick"].values)].index.to_list()
    #     if len(obs_row_index) == 0:
    #         return False
    #     else:
    #         return True
    #
    # def approximate_suggestion(self, x_opt, grid_unscaled):
    #
    #     l2_penalty = 0
    #     searchgrid_p1_matched_rows = grid_unscaled[(grid_unscaled["p1_mat"] == x_opt['p1_mat'].to_numpy()[0])]
    #     searchgrid_p2_matched_rows = grid_unscaled[(grid_unscaled["p2_mat"] == x_opt['p2_mat'].to_numpy()[0])]
    #     searchgrid_p3_matched_rows = grid_unscaled[(grid_unscaled["p3_mat"] == x_opt['p3_mat'].to_numpy()[0])]
    #
    #     if searchgrid_p1_matched_rows.empty:
    #         l2_penalty += 1
    #     if searchgrid_p2_matched_rows.empty:
    #         l2_penalty += 1
    #     if searchgrid_p3_matched_rows.empty:
    #         l2_penalty += 1
    #
    #     searchgrid_p1p2p3_matched_rows = grid_unscaled[(grid_unscaled["p1_mat"] == x_opt['p1_mat'].to_numpy()[0]) &
    #                                              (grid_unscaled["p2_mat"] == x_opt['p2_mat'].to_numpy()[0]) &
    #                                              (grid_unscaled["p3_mat"] == x_opt['p3_mat'].to_numpy()[0])]
    #
    #     if not searchgrid_p1p2p3_matched_rows.empty:
    #
    #         min_values_np = grid_unscaled[self.exp_params.l2norm_colnames].min().to_numpy()
    #         max_values_np = grid_unscaled[self.exp_params.l2norm_colnames].max().to_numpy()
    #         grid_obs_thick_gap = searchgrid_p1p2p3_matched_rows[self.exp_params.l2norm_colnames]
    #         grid_obs_thick_gap_np = grid_obs_thick_gap.to_numpy()
    #         grid_obs_thick_gap_np_norm = (grid_obs_thick_gap_np - min_values_np)/(max_values_np - min_values_np)
    #         xopt_thick_gap = x_opt[self.exp_params.l2norm_colnames]
    #         xopt_thick_gap_np = xopt_thick_gap.to_numpy()
    #         xopt_thick_gap_np_norm = (xopt_thick_gap_np - min_values_np)/(max_values_np - min_values_np)
    #         difference = grid_obs_thick_gap_np_norm - xopt_thick_gap_np_norm
    #         l2_norm = np.linalg.norm(difference, axis=1)
    #         l2_norm += l2_penalty
    #
    #         count = 1
    #         while count <= self.exp_params.approx_steps:
    #             min_norm_index = np.argmin(l2_norm)
    #             min_norm = l2_norm[min_norm_index]
    #             bo_suggestion = searchgrid_p1p2p3_matched_rows.iloc[min_norm_index, 1:]
    #             bo_suggestion_pd = pd.DataFrame(bo_suggestion).T
    #             if not self.check_if_duplicates(bo_suggestion_pd):
    #                 return bo_suggestion_pd, min_norm, min_norm_index
    #             else:
    #                 l2_norm[min_norm_index] = np.inf
    #                 print(colored("Repeated suggestion found", "yellow"))
    #             count += 1
    #
    #         if count > self.exp_params.approx_steps:
    #             print(colored("Couldn't approximate suggestion, given the steps, please try updating manually", "red"))
    #             return None, None, None
    #
    # def generate_grid_file(self):
    #     X = np.array([i for i in range(1, self.exp_params.num_samples_grid+1)])
    #     weight_data = self.weight_function(X)
    #     penetration_data = self.penetration_function(X)
    #     df = pd.DataFrame(columns=self.exp_params.colnames_grid)
    #     df["ID"] = X
    #     df["Weight"] = weight_data
    #     df["Penetration"] = penetration_data
    #     df.to_excel(self.exp_params.grid_filename+".xlsx", index=False)
    #     df.to_pickle(self.exp_params.grid_filename+".pkl")
    #
    # def update_files_results(self, bo_suggestion, itr, min_norm=None):
    #
    #     bo_suggestion.insert(0, str(itr + self.exp_params.num_init_obs))
    #     bo_suggestion.append("L2-Norm: " + str(min_norm))
    #
    #     current_obsfile_mod_time = os.path.getmtime(self.exp_params.obs_file)
    #     wb = openpyxl.load_workbook(filename=self.exp_params.obs_file)
    #     sheet = wb[self.exp_params.obs_sheetname]
    #     sheet.append(bo_suggestion)
    #     wb.save(self.exp_params.obs_file)
    #     wb.close()
    #     # #Reset file (accesstime, modificationtime) to not let the wrapper think expert updated the results
    #     os.utime(self.exp_params.obs_file, (current_obsfile_mod_time, current_obsfile_mod_time))
    #     time.sleep(1)
